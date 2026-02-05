import re
import requests
import openpyxl
from pathlib import Path
from urllib.parse import urlparse
import numpy as np


XLSX_PATH = r"C:\Users\p2515497\Documents\WIEN_CONF_2025\Abstract_list.xlsx"
OUT_TEX   = r"C:\Users\p2515497\Documents\WIEN_CONF_2025\BookAbstract.tex"
SHEET_NAME = None		  # e.g. "Sheet1" or None for active sheet

START_ROW = 3
URL_COL = 3				# column C
TIMEOUT = 60

AREA_ORDER  = ["Plenary","Damage Mechanics","Optimization & Dynamic Response", "Delamination & Impact", "Novel Approaches", "Fracture Mechanics","Thin Ply", "Buckling / Stability","Structures","Multi-scale modeling","Novel Materials","Machine Learning I","Machine Learning II"]

SCALE = 0.95

AREA_COL = 9

TITLE_COL = 2
AUTHOR_COL = 10
	
	
# ------------------------- SETTINGS (EDIT THESE) -------------------------
_LATEX_SPECIALS = {
    "\\": r"\textbackslash{}",
    "&": r"\&",
    "%": r"\%",
    "$": r"\$",
    "#": r"\#",
    "_": r"\_",
    "{": r"\{",
    "}": r"\}",
    "~": r"\textasciitilde{}",
    "^": r"\textasciicircum{}",
}

# -----------------------------------------------------------------------


#%% Function for handling the excel file 
def sanitize_filename(name: str, max_len: int = 80) -> str:
	name = (name or "").strip()					# remove all the whitespace from the start and end of the file name string
	name = re.sub(r"\s+", " ", name)			# replace all thewhitespace by regular space space
	name = re.sub(r'[\\/:*?"<>|]+', "", name)  	# remove forbidden character
	name = name.strip(" .")
	if not name:
		name = "document"
	return name[:max_len]

def get_cell_text(cell) -> str:
	"""
	Get the text content of an excel cell
	"""
	if cell is None or cell.value is None:
		return ""
	return str(cell.value).strip()

def get_cell_url(cell) -> str | None:
	"""
	Get the url content of an excel cell
	"""
	# Excel hyperlink target has priority
	if cell.hyperlink and cell.hyperlink.target:
		return str(cell.hyperlink.target).strip()
	# Otherwise, cell text
	val = get_cell_text(cell)
	if val.startswith("http://") or val.startswith("https://"):
		return val
	return None

def parse_main_author(authors_cell: str) -> str:
	"""
	From the list of all authors involved in an abstract, return the main author, which has been identified in the
	Excel cell by the symbol "*"

	Parameters
	----------
	authors_cell : str
		cell containing the list of authors.

	Returns
	-------
	str
		Name of the main author.

	"""
	authors = [a.strip() for a in (authors_cell or "").split(",") if a.strip()]
	if not authors:
		return "Unknown"
	main = None
	for a in authors:
		if "*" in a:
			main = a
			break
	if main is None:
		main = authors[0]
	# remove the star marker and extra spaces
	main = main.replace("*", "").strip()
	return main if main else "Unknown"

def parse_all_authors(authors_cell: str) -> list[str]:
	"""
	From the list of all authors involved in an abstract, return the a list containing all the involve author

	Parameters
	----------
	authors_cell : str
		cell containing the list of authors.

	Returns
	-------
	list[str]
		List of all authors involved in an abastract.

	"""
	# comma-separated; strip whitespace; remove '*' marker
	authors = [a.strip() for a in (authors_cell or "").split(",") if a.strip()]
	clean = []
	for a in authors:
		a = a.replace("*", "").strip()
		if a:
			clean.append(a)
	return clean


#%% File helper functions 
def filename_from_url(url: str, idx: int) -> str:
	parsed = urlparse(url)						# Takes a URL string and splits it into its parts
	name = Path(parsed.path).name 				# Extract the url pdf name
	if not name.lower().endswith(".pdf") or not name:   # if at the end of the url there is no name, a default file name is assgined
		name = f"file_{idx:04d}.pdf"
	return sanitize_filename(name)

def is_pdf_file(path: Path) -> bool:
	"""
	Check if the file is a pdf or not. Pay attention, the variable is called path but it is actually just the file name
	Try to retrun the first 5 bytes of the file.
	If it is a PDF, return the 5 first bytes, else return False
	Parameters
	----------
	path : Path object
		file name.

	Returns
	-------
	bool
		Boolean to know if the url links to a pdf or not.

	"""
	# 
	# 
	try:
		with open(path, "rb") as f:
			return f.read(5) == b"%PDF-"
	except Exception:
		return False

def download_pdf(url: str, out_path: Path, timeout: int = TIMEOUT) -> None:
	"""
	

	Parameters
	----------
	url : str
		url link to download the pdf from.
	out_path : Path
		path where to store locally the pdf .
	timeout : int, optional
		waiting time before giving up. The default is TIMEOUT.

	Returns
	-------
	None
		DESCRIPTION.

	"""
	headers = {"User-Agent": "xlsx-pdf-embed/1.0"} 									#Extra care to make the request looks "normal" and is not blocked by the website (might be useless
	with requests.get(url, headers=headers, stream=True, timeout=timeout) as r:
		r.raise_for_status() 										# Raise an error if there is an error, instead of downloading the error message
		out_path.parent.mkdir(parents=True, exist_ok=True) 			# Create the directory to store locally the pdf
		with open(out_path, "wb") as f: 							# Download the pdf and store them in the right path
			for chunk in r.iter_content(chunk_size=1024 * 128):
				if chunk:
					f.write(chunk)
	return





#%% Latex related functions 

def build_tex(records: list[dict], out_tex: Path) -> None:
	"""
	records list items contain:
	- area, title, main_author, pdf_path, id
	Already filtered to AREA_ORDER and ordered by AREA_ORDER then Excel row order.
	"""
	parts = []
	parts.append(r"\documentclass[11pt]{article}")
	parts.append(r"\usepackage[a4paper,margin=1.5cm]{geometry}")
	parts.append(r"\usepackage{pdfpages}")
	parts.append(r"\usepackage{hyperref}")
	parts.append(r"\hypersetup{hidelinks}")
	parts.append(r"\usepackage{tabularx}")
	parts.append(r"\usepackage{ltablex}")
	parts.append(r"\usepackage{longtable}")
	parts.append(r"\usepackage{array}")
	parts.append(r"\usepackage{fontspec}")
	parts.append(r"\usepackage{xcolor}")
	parts.append(r"\definecolor{HeaderGreen}{RGB}{0,120,60}")  
	parts.append(r"\setlength{\headheight}{24pt}")             
	parts.append(r"\setmainfont{Arial Narrow}")
	parts.append(r"\keepXColumns")
	parts.append(r"\usepackage{fancyhdr}")
	parts.append(r"\pagestyle{fancy}")
	parts.append(r"\fancyhf{}")
	parts.append(r"\fancyhead{}")
	parts.append(
		r"\fancyhead[C]{"
		r"\colorbox{HeaderGreen}{"
		r"\parbox{\textwidth}{\centering\color{white}\large COMPOSITE 2025 - Vienna - Book of abstracts}"
		r"}"
		r"}"
	)
	parts.append(r"\renewcommand{\headrulewidth}{0pt}")  # Linie aus, weil der Balken die „Linie“ ist
	parts.append(r"\renewcommand{\headrulewidth}{0.4pt}")
	parts.append(r"\fancyfoot[C]{\thepage}")
	parts.append(r"\renewcommand{\headrulewidth}{0pt}")
	parts.append(r"\newcommand{\CurrentPDFTarget}{}")
	parts.append(r"\newcommand{\CurrentPDFLabel}{}")


	parts.append(r"\begin{document}")
	parts.append(r"\includepdf[pages=1-4,scale=1,pagecommand={\thispagestyle{empty}}]{Intro.pdf}")


	# Custom “TOC” before everything
	parts.extend(make_custom_toc(records))

	current_area = None

	for rec in records:
		area = rec["area"]
		if area != current_area:
			current_area = area
			parts.extend(make_transition_page(area))

		pdf_path = rec["pdf_path"]
		target_id = rec["id"]
		latex_path = str(pdf_path).replace("\\", "/")

		# Set a one-shot target name; pagecommand will place it on the FIRST included page and then clear it
		parts.append(r"\gdef\CurrentPDFTarget{%s}" % target_id)


		target_id = rec["id"]              # e.g. abs:0001
		label_id  = rec["label"]           # e.g. lab:0001
		
		parts.append(r"\gdef\CurrentPDFTarget{%s}" % rec["id"])
		parts.append(r"\gdef\CurrentPDFLabel{%s}" % rec["label"])
		if SCALE != 1.0:
			parts.append(
				r"\includepdf[pages=-,scale=%s,pagecommand={\thispagestyle{fancy}"
				r"\ifx\CurrentPDFTarget\empty\else"
				r"\phantomsection"
				r"\hypertarget{\CurrentPDFTarget}{}"
				r"\label{\CurrentPDFLabel}"
				r"\gdef\CurrentPDFTarget{}\gdef\CurrentPDFLabel{}"
				r"\fi}]{%s}" % (SCALE, latex_path)
			)
		else:
			parts.append(
				r"\includepdf[pages=-,pagecommand={\thispagestyle{fancy}"
				r"\ifx\CurrentPDFTarget\empty\else"
				r"\phantomsection"
				r"\hypertarget{\CurrentPDFTarget}{}"
				r"\label{\CurrentPDFLabel}"
				r"\gdef\CurrentPDFTarget{}\gdef\CurrentPDFLabel{}"
				r"\fi}]{%s}" % latex_path
			)
	
	
	author_index = build_author_index(records)
	parts.extend(make_author_index_section(author_index))
	
	
	parts.append(r"\end{document}")

	out_tex.parent.mkdir(parents=True, exist_ok=True)
	out_tex.write_text("\n".join(parts), encoding="utf-8")
	print(f"Wrote LaTeX file: {out_tex}")

	return
def make_custom_toc(entries: list[dict]) -> list[str]:
	"""
	entries: list of dicts with keys:
	- area
	- id
	- label
	- main_author
	- title

	Produces a multi-page TOC grouped by AREA_ORDER.
	"""
	parts = []
	parts.append(r"\pagestyle{empty}")
	parts.append(r"\begin{center}")
	parts.append(r"\LARGE Table of Contents")
	parts.append(r"\end{center}")
	parts.append(r"\vspace{1em}")

	# More spacing between rows
	parts.append(r"\renewcommand{\arraystretch}{1.35}")
	parts.append(r"\setlength{\tabcolsep}{6pt}")

	# 3 columns: Author | Title | Page
	parts.append(r"\noindent\begin{longtable}{@{}>{\bfseries}p{0.28\textwidth} p{0.62\textwidth} r@{}}")

	# (Optional) header row repeated on page breaks
	parts.append(r"\textbf{Author} & \textbf{Title} & \textbf{Page}\\")
	parts.append(r"\hline")
	parts.append(r"\endfirsthead")
	parts.append(r"\textbf{Author} & \textbf{Title} & \textbf{Page}\\")
	parts.append(r"\hline")
	parts.append(r"\endhead")

	# Group entries by area (preserve order within each area)
	by_area: dict[str, list[dict]] = {a: [] for a in AREA_ORDER}
	for e in entries:
		a = e.get("area", "")
		if a in by_area:
			by_area[a].append(e)

	for area in AREA_ORDER:
		if not by_area[area]:
			continue

		# Session header row spanning all 3 columns
		area_tex = latex_escape(area)
		parts.append(r"\multicolumn{3}{@{}l@{}}{\Large\bfseries %s}\\[4pt]" % area_tex)
		parts.append(r"\hline")
		parts.append(r"\noalign{\vskip 4pt}")
		# Entries for that session
		for e in by_area[area]:
			ma = latex_escape(e["main_author"])
			ti = latex_escape(e["title"])
			link = r"\hyperlink{%s}{%s}" % (e["id"], ti)
			page = r"\pageref{%s}" % e["label"]

			parts.append(r"%s & %s & %s \\" % (ma, link, page))
			parts.append(r"\noalign{\vskip 3pt}")  # extra spacing between entries

		# Extra space between sessions
		parts.append(r"\noalign{\vskip 8pt}")

	parts.append(r"\end{longtable}")
	parts.append(r"\clearpage")
	parts.append(r"\pagestyle{fancy}")
	return parts


def make_author_index_section(author_index: dict[str, list[dict]]) -> list[str]:
	"""
	author_index: { "Author Name": [ {id,label,title,area}, ... ], ... }

	Outputs a 3-column Author Index:
	- Col 1: initials/prefix (right-aligned)
	- Col 2: surname (left-aligned, all start same x)
	- Col 3: pages (clickable numbers linking to abstracts)
	"""
	parts = []
	parts.append(r"\clearpage")
	parts.append(r"\section*{Author Index}")
	parts.append(r"\addcontentsline{toc}{section}{Author Index}")
	parts.append(r"\vspace{0.5em}")

	# spacing
	parts.append(r"\renewcommand{\arraystretch}{1.2}")
	parts.append(r"\setlength{\tabcolsep}{6pt}")

	# 3 columns: Initials | Surname | Pages
	# - initials: narrow fixed column, right aligned
	# - surname: wider fixed column, left aligned
	# - pages: remaining space
	parts.append(r"\noindent\begin{longtable}{@{}p{0.1\textwidth} p{0.42\textwidth} p{0.46\textwidth}@{}}")

	# header (repeats on page breaks)
	parts.append(r"\textbf{Initials} & \textbf{Surname} & \textbf{Pages}\\")
	parts.append(r"\hline")
	parts.append(r"\endfirsthead")
	parts.append(r"\textbf{Initials} & \textbf{Surname} & \textbf{Pages}\\")
	parts.append(r"\hline")
	parts.append(r"\endhead")

	authors_sorted = sorted(author_index.keys(), key=author_sort_key)

	for author in authors_sorted:
		items = author_index.get(author, [])

		# keep pages in document order
		items_sorted = sorted(items, key=lambda it: it.get("id", ""))

		# clickable page list: click number -> jump to abstract anchor
		links = []
		for it in items_sorted:
			links.append(r"\hyperlink{%s}{\pageref{%s}}" % (it["id"], it["label"]))
		pages_tex = ", ".join(links) if links else ""

		initials, surname = split_author_initials_surname(author)
		initials_tex = latex_escape(initials)
		surname_tex = latex_escape(surname)

		# Make initials visually right-aligned within their column
		parts.append(r"\makebox[\linewidth][r]{%s} & %s & %s \\" % (initials_tex, surname_tex, pages_tex))
		parts.append(r"\noalign{\vskip 2pt}")

	parts.append(r"\end{longtable}")
	return parts


def build_author_index(records: list[dict]) -> dict[str, list[dict]]:
	"""
	Returns: { "Author Name": [ {id, label, title, area}, ... ], ... }
	"""
	idx: dict[str, list[dict]] = {}
	for r in records:
		for a in r.get("authors", []):
			idx.setdefault(a, []).append({
				"id": r["id"],
				"label": r["label"],
				"title": r["title"],
				"area": r["area"],
			})
	# sort authors alphabetically (case-insensitive)
	sorted_items = sorted(idx.items(), key=lambda kv: kv[0].casefold())
	return dict(sorted(idx.items(), key=lambda kv: author_sort_key(kv[0])))


def latex_escape(text: str) -> str:
    """Escape LaTeX special characters, but keep normal Unicode."""
    if text is None:
        return ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return "".join(_LATEX_SPECIALS.get(ch, ch) for ch in text)


def make_transition_page(area: str) -> list[str]:
	# A clean title/transition page; no page number shown
	area_esc = latex_escape(area)
	return [
		r"\clearpage",
		r"\thispagestyle{empty}",
		r"\vspace*{\fill}",
		r"\begin{center}",
		r"\Huge " + area_esc,
		r"\end{center}",
		r"\vspace*{\fill}",
		r"\clearpage",
	]

	return



	






def author_sort_key(name: str) -> tuple[str, str]:
	initials, surname = split_author_initials_surname(name)
	return (surname.casefold(), initials.casefold(), name.casefold())


def split_author_initials_surname(name: str) -> tuple[str, str]:
	"""
	Split at the end of the initials (first '.' sequence).
	Examples:
	- 'P. Van Morisson' -> ('P.', 'Van Morisson')
	- 'P.A. Van Morisson' -> ('P.A.', 'Van Morisson')
	- 'P. A. Van Morisson' -> ('P. A.', 'Van Morisson')
	"""
	s = (name or "").strip()
	s = re.sub(r"\s+", " ", s)

	# If there is no dot, fallback: no initials, whole string is surname/name
	if "." not in s:
		return ("", s)

	# Find the position of the LAST dot within the "initials block" at the start.
	# This matches patterns like: "P.", "P.A.", "P. A.", "P. A. B."
	m = re.match(r"^([A-Za-z](?:\.\s*[A-Za-z])*\.)\s*(.*)$", s)
	if m:
		initials = m.group(1).replace(" ", "") if ". " not in m.group(1) else m.group(1)
		rest = m.group(2).strip()
		return (initials.strip(), rest if rest else "")

	# Simpler fallback: split at first '.' only
	i = s.find(".")
	return (s[:i+1].strip(), s[i+1:].strip())








def main() -> None:
	xlsx_path = Path(XLSX_PATH)
	out_tex = Path(OUT_TEX)

	wb = openpyxl.load_workbook(xlsx_path, data_only=True)
	ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

	pdf_dir = out_tex.parent / "downloaded_pdfs"
	per_area: dict[str, list[dict]] = {a: [] for a in AREA_ORDER}

	# 1) READ EXCEL ONCE
	for row in range(START_ROW, ws.max_row + 1):
		url = get_cell_url(ws.cell(row=row, column=URL_COL))
		if not url:
			continue

		area = get_cell_text(ws.cell(row=row, column=AREA_COL))
		if area not in AREA_ORDER:
			continue

		title = get_cell_text(ws.cell(row=row, column=TITLE_COL))
		authors_cell = get_cell_text(ws.cell(row=row, column=AUTHOR_COL))
		main_author = parse_main_author(authors_cell)
		authors_list = parse_all_authors(authors_cell)

		per_area[area].append({
			"row": row,
			"area": area,
			"title": title if title else "(No title)",
			"main_author": main_author,
			"authors": authors_list,
			"url": url,		})

	# 2) DOWNLOAD EACH PDF ONCE (IN AREA_ORDER)
	records: list[dict] = []
	idx = 0

	for area in AREA_ORDER:
		for rec in per_area[area]:
			idx += 1
			local_name = filename_from_url(rec["url"], idx)
			pdf_path = pdf_dir / local_name

			try:
				# Optional: skip download if already present and valid
				if pdf_path.exists() and is_pdf_file(pdf_path):
					pass
				else:
					download_pdf(rec["url"], pdf_path)

				if not is_pdf_file(pdf_path):
					print(f"NOT A PDF (skipping) row {rec['row']}: {pdf_path.name}")
					try:
						pdf_path.unlink()
					except Exception:
						pass
					continue

				rec2 = dict(rec)
				rec2["pdf_path"] 	= pdf_path
				rec2["id"] 			= f"abs:{idx:04d}"	# hyperlink id
				rec2["label"] 		= f"lab:{idx:04d}"
				records.append(rec2)

				print(f"OK row {rec['row']} | area={area} | file={pdf_path.name}")

			except Exception as e:
				print(f"ERROR row {rec['row']} area={area} url={rec['url']} reason={e}")

	# 3) BUILD TEX ONCE
	if not records:
		raise RuntimeError("No valid PDFs downloaded for the requested AREA_ORDER list.")

	build_tex(records, out_tex)
	print(f"PDFs saved under: {pdf_dir}")
	print("Compile from the folder containing the .tex:")
	print(f"  pdflatex {out_tex.name}")
	print("If pdflatex fails due to PDF compatibility, try:")
	print(f"  lualatex {out_tex.name}")


if __name__ == "__main__":
	main()
