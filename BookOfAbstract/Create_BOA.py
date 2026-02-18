#%% Libraries

import re
import requests
import openpyxl
from pathlib import Path
from urllib.parse import urlparse
import numpy as np
import shutil


#%% Variables

#%%% Directory

# Path to the excel file
XLSX_PATH 			= Path(r"Z:\ikkv\Dokus_LKKV\040_Projekte\02_Eigenforschung\2025-09-Thomas\Book-of-Abstracts/Abstract_list.xlsx")

# Path where to store the .tex file
OUT_TEX   			= Path(r"Z:\ikkv\Dokus_LKKV\040_Projekte\02_Eigenforschung\2025-09-Thomas\Book-of-Abstracts/BookAbstract.tex")

# If no PDF is found following the URL, look for the PDF in the following folder
LOCAL_FALLBACK_DIR = Path(r"C:\Users\p2515497\Documents\WIEN_CONF_2025\LOCAL_PDFS")


USE_URL_CELL_AS_LOCAL_FILENAME = True


#%%% Excel variables
SHEET_NAME 		= None		  # e.g. "Sheet1" or None for active sheet

START_ROW 		= 3
URL_COL 		= 3				# column C
AREA_COL 		= 9
TITLE_COL 		= 2
AUTHOR_COL 		= 10
TIMEOUT 		= 60

AREA_ORDER  	= ["Plenary","Damage Mechanics","Optimization & Dynamic Response", "Delamination & Impact", "Novel Approaches", "Fracture Mechanics","Thin Ply", "Buckling / Stability","Structures","Multi-scale modeling","Novel Materials","Machine Learning I","Machine Learning II"]


#%%% Tex variable

# scaling each pdf, must be <1 to insert correctly a header and a page numbering in the tex file
SCALE 			= 0.95

_LATEX_SPECIALS = {
		"\\": r"\textbackslash{}",
		"&": r"\&",
		"%": r"\%",
		"$": r"\$",
		"#": r"\#",
		"_": r"\_",
		"{": r"\{",
		"}": r"\}",
		"~": r"\textasciitilde{}",
		"^": r"\textasciicircum{}",
}

# -----------------------------------------------------------------------


#%% Function for handling the excel file 
def sanitize_filename(name: str, max_len: int = 80) -> str:
	name = (name or "").strip()					# remove all the whitespace from the start and end of the file name string
	name = re.sub(r"\s+", " ", name)			# replace all thewhitespace by regular space space
	name = re.sub(r'[\\/:*?"<>|]+', "", name)  	# remove forbidden character
	name = name.strip(" .")
	if not name:
		name = "document"
	return name[:max_len]

def get_cell_text(cell) -> str:
	"""
	Get the text content of an excel cell
	"""
	if cell is None or cell.value is None:
		return ""
	return str(cell.value).strip()

def get_cell_url(cell) -> str | None:
	"""
	Get the url content of an excel cell
	"""
	# Excel hyperlink target has priority
	if cell.hyperlink and cell.hyperlink.target:
		return str(cell.hyperlink.target).strip()
	# Otherwise, cell text
	val = get_cell_text(cell)
	if val.startswith("http://") or val.startswith("https://"):
		return val
	return None

def parse_main_author(authors_cell: str) -> str:
	"""
	From the list of all authors involved in an abstract, return the main author, which has been identified in the
	Excel cell by the symbol "*"

	Parameters
	----------
	authors_cell : str
		cell containing the list of authors.

	Returns
	-------
	str
		Name of the main author.

	"""
	authors = [a.strip() for a in (authors_cell or "").split(",") if a.strip()]
	if not authors:
		return "Unknown"
	main = None
	for a in authors:
		if "*" in a:
			main = a
			break
	if main is None:
		main = authors[0]
	# remove the star marker and extra spaces
	main = main.replace("*", "").strip()
	return main if main else "Unknown"

def parse_all_authors(authors_cell: str) -> list[str]:
	"""
	From the list of all authors involved in an abstract, return the a list containing all the involve author

	Parameters
	----------
	authors_cell : str
		cell containing the list of authors.

	Returns
	-------
	list[str]
		List of all authors involved in an abastract.

	"""
	# comma-separated; strip whitespace; remove '*' marker
	authors = [a.strip() for a in (authors_cell or "").split(",") if a.strip()]
	clean = []
	for a in authors:
		a = a.replace("*", "").strip()
		if a:
			clean.append(a)
	return clean


#%% File helper functions 

def find_local_pdf(rec: dict, idx: int, url_cell_text: str = "") -> Path | None:
	"""
	Findet eine lokale PDF im LOCAL_FALLBACK_DIR, wenn keine URL vorhanden ist.
	Strategie:
	  1) Falls in Excel-URL-Zelle ein Dateiname steht -> exakt suchen
	  2) Sonst: sanitize(title)+.pdf -> exakt suchen
	  3) Sonst: heuristische Suche per glob
	"""
	base = LOCAL_FALLBACK_DIR
	if not base.exists():
			print(f"LOCAL_FALLBACK_DIR existiert nicht: {base}")
			return None

	candidates = []

	if USE_URL_CELL_AS_LOCAL_FILENAME:
			txt = (url_cell_text or "").strip()
			if txt and (txt.lower().endswith(".pdf") or "." in txt) and ("http://" not in txt and "https://" not in txt):
					candidates.append(base / txt)

	# 2) Default: Titel als Dateiname
	title = rec.get("title", "")
	if title:
			candidates.append(base / (sanitize_filename(title) + ".pdf"))

	# 3) Fallback: nach ähnlichen PDFs suchen
	#		(z.B. erster längerer Token aus Titel)
	tokens = [t for t in re.split(r"\W+", title) if len(t) >= 6]
	if tokens:
			pat = f"*{tokens[0]}*.pdf"
			hits = list(base.glob(pat))
			if hits:
					# nimm den ersten Treffer
					candidates.append(hits[0])

	# Check candidates
	for p in candidates:
			try:
					if p.exists() and is_pdf_file(p):
							return p
			except Exception:
					pass

	return None




def filename_from_url(url: str, idx: int) -> str:
	parsed = urlparse(url)
	name = Path(parsed.path).name
	if not name.lower().endswith(".pdf") or not name:
		name = "file.pdf"
	stem = Path(name).stem
	return sanitize_filename(f"{stem}_{idx:04d}.pdf")


def filename_for_record(rec: dict, idx: int) -> str:
	"""
	Return a stable local filename for this record.
	- If URL exists: derive from URL (original behavior)
	- Else: derive from local filename in url_text (if present), else from title, else fallback on idx.
	"""
	url = rec.get("url")
	if url:
		return filename_from_url(url, idx)

	# no URL -> try local filename from the URL cell text
	url_text = (rec.get("url_text") or "").strip()
	if url_text:
		# if user wrote something like "myfile.pdf" in the URL cell
		name = Path(url_text).name
		if name:
			if not name.lower().endswith(".pdf"):
				name += ".pdf"
			return sanitize_filename(name)

	# else build from title
	title = (rec.get("title") or "").strip()
	if title:
		return sanitize_filename(title) + ".pdf"

	# final fallback
	return f"local_{idx:04d}.pdf"



def is_pdf_file(path: Path) -> bool:
	"""
	Check if the file is a pdf or not. Pay attention, the variable is called path but it is actually just the file name
	Try to retrun the first 5 bytes of the file.
	If it is a PDF, return the 5 first bytes, else return False
	Parameters
	----------
	path : Path object
		file name.

	Returns
	-------
	bool
		Boolean to know if the url links to a pdf or not.

	"""
	# 
	# 
	try:
		with open(path, "rb") as f:
			return f.read(5) == b"%PDF-"
	except Exception:
		return False

def download_pdf(url: str, out_path: Path, timeout: int = TIMEOUT) -> None:
	"""
	

	Parameters
	----------
	url : str
		url link to download the pdf from.
	out_path : Path
		path where to store locally the pdf .
	timeout : int, optional
		waiting time before giving up. The default is TIMEOUT.

	Returns
	-------
	None
		DESCRIPTION.

	"""
	headers = {"User-Agent": "xlsx-pdf-embed/1.0"} 									#Extra care to make the request looks "normal" and is not blocked by the website (might be useless
	with requests.get(url, headers=headers, stream=True, timeout=timeout) as r:
		r.raise_for_status() 										# Raise an error if there is an error, instead of downloading the error message
		out_path.parent.mkdir(parents=True, exist_ok=True) 			# Create the directory to store locally the pdf
		with open(out_path, "wb") as f: 							# Download the pdf and store them in the right path
			for chunk in r.iter_content(chunk_size=1024 * 128):
				if chunk:
					f.write(chunk)
	return





def build_tex(records: list[dict], out_tex: Path) -> None:
	parts = []
	parts.append(r"\documentclass[11pt]{article}")
	parts.append(r"\usepackage[a4paper,margin=1.5cm]{geometry}")
	parts.append(r"\setlength{\footskip}{1mm}")  # try 6mm–12mm
	parts.append(r"\usepackage{pdfpages}")
	parts.append(r"\usepackage{hyperref}")
	parts.append(r"\hypersetup{hidelinks}")
	parts.append(r"\usepackage{tabularx}")
	parts.append(r"\usepackage{ltablex}")
	parts.append(r"\usepackage{longtable}")
	parts.append(r"\usepackage{array}")
	parts.append(r"\usepackage{fontspec}")
	parts.append(r"\usepackage{xcolor}")
	parts.append(r"\usepackage{multicol}")
	parts.append(r"\definecolor{HeaderGreen}{RGB}{0,120,60}")  
	parts.append(r"\setlength{\headheight}{24pt}")						 
	parts.append(r"\setmainfont{DINish}")
	parts.append(r"\keepXColumns")
	parts.append(r"\usepackage{fancyhdr}")
	parts.append(r"\pagestyle{fancy}")
	parts.append(r"\fancyhf{}")
	parts.append(r"\fancyhead{}")
	parts.append(
	    r"\fancyhead[R]{"
	    r"\footnotesize "
	    r"10th ECCOMAS Thematic Conference on the Mechanical Response of Composites: COMPOSITES 2025\\"
	    r"H.E. \textbf{Pettermann}, C. \textbf{Schuecker}, M. \textbf{Fagerström} (Eds)"
	    r"}"
	)
	parts.append(r"\renewcommand{\headrulewidth}{0pt}")  # Linie aus, weil der Balken die „Linie“ ist
	parts.append(r"\renewcommand{\headrulewidth}{0.4pt}")
	parts.append(r"\fancyfoot[C]{\thepage}")
	parts.append(r"\renewcommand{\headrulewidth}{0pt}")
	parts.append(r"\newcommand{\CurrentPDFTarget}{}")
	parts.append(r"\newcommand{\CurrentPDFLabel}{}")


	parts.append(r"\begin{document}")
	parts.append(r"\includepdf[pages=1-4,scale=1,pagecommand={\thispagestyle{empty}}]{Book-of-Abstracts_Front-part.pdf}")
	parts.append(r"\pagenumbering{arabic}")
	parts.append(r"\setcounter{page}{1}")
	parts.extend(make_custom_toc(records))

	current_area = None

	for rec in records:
		area = rec["area"]
		if area != current_area:
			current_area = area
			parts.extend(make_transition_page(area))

		pdf_path = rec["pdf_path"]
		latex_path = str(pdf_path).replace("\\", "/")

		# set target + label ONCE
		parts.append(r"\gdef\CurrentPDFTarget{%s}" % rec["id"])
		parts.append(r"\gdef\CurrentPDFLabel{%s}" % rec["label"])

		if SCALE != 1.0:
			parts.append(
				r"\includepdf[pages=-,scale=%s,pagecommand={\thispagestyle{fancy}"
				r"\ifx\CurrentPDFTarget\empty\else"
				r"\phantomsection"
				r"\hypertarget{\CurrentPDFTarget}{}"
				r"\label{\CurrentPDFLabel}"
				r"\gdef\CurrentPDFTarget{}\gdef\CurrentPDFLabel{}"
				r"\fi}]{%s}" % (SCALE, latex_path)
			)
		else:
			parts.append(
				r"\includepdf[pages=-,pagecommand={\thispagestyle{fancy}"
				r"\ifx\CurrentPDFTarget\empty\else"
				r"\phantomsection"
				r"\hypertarget{\CurrentPDFTarget}{}"
				r"\label{\CurrentPDFLabel}"
				r"\gdef\CurrentPDFTarget{}\gdef\CurrentPDFLabel{}"
				r"\fi}]{%s}" % latex_path
			)

	# ---- AUTHOR INDEX ONCE (outside loop) ----
	author_index = build_author_index(records)
	parts.extend(make_author_index_section(author_index))
	
	parts.append(r"\includepdf[pages=1,scale=1,pagecommand={\thispagestyle{empty}}]{Book-of-Abstracts_final-page.pdf}")
	parts.append(r"\end{document}")

	out_tex.parent.mkdir(parents=True, exist_ok=True)
	out_tex.write_text("\n".join(parts), encoding="utf-8")
	print(f"Wrote LaTeX file: {out_tex}")






def make_custom_toc(entries: list[dict]) -> list[str]:
	"""
	entries: list of dicts with keys:
	- area
	- id
	- label
	- main_author
	- title

	Produces a multi-page TOC grouped by AREA_ORDER.
	"""
	parts = []
	parts.append(r"\pagestyle{empty}")
	parts.append(r"\begin{center}")
	parts.append(r"\LARGE Table of Contents")
	parts.append(r"\end{center}")
	parts.append(r"\vspace{1em}")

	# More spacing between rows
	parts.append(r"\renewcommand{\arraystretch}{1.35}")
	parts.append(r"\setlength{\tabcolsep}{6pt}")

	# 3 columns: Author | Title | Page
	parts.append(r"\noindent\begin{longtable}{@{}>{\bfseries}p{0.28\textwidth} p{0.62\textwidth} r@{}}")

	# (Optional) header row repeated on page breaks
	parts.append(r"\textbf{Author} & \textbf{Title} & \textbf{Page}\\")
	parts.append(r"\hline")
	parts.append(r"\endfirsthead")
	parts.append(r"\textbf{Author} & \textbf{Title} & \textbf{Page}\\")
	parts.append(r"\hline")
	parts.append(r"\endhead")

	# Group entries by area (preserve order within each area)
	by_area: dict[str, list[dict]] = {a: [] for a in AREA_ORDER}
	for e in entries:
		a = e.get("area", "")
		if a in by_area:
			by_area[a].append(e)

	for area in AREA_ORDER:
		if not by_area[area]:
			continue

		# Session header row spanning all 3 columns
		area_tex = latex_escape(area)
		parts.append(r"\multicolumn{3}{@{}l@{}}{\Large\bfseries %s}\\[4pt]" % area_tex)
		parts.append(r"\hline")
		parts.append(r"\noalign{\vskip 4pt}")
		# Entries for that session
		for e in by_area[area]:
			ma = latex_escape(e["main_author"])
			ti = latex_escape(e["title"])
			link = r"\hyperlink{%s}{%s}" % (e["id"], ti)
			page = r"\pageref{%s}" % e["label"]

			parts.append(r"%s & %s & %s \\" % (ma, link, page))
			parts.append(r"\noalign{\vskip 3pt}")  # extra spacing between entries

		# Extra space between sessions
		parts.append(r"\noalign{\vskip 8pt}")

	parts.append(r"\end{longtable}")
	parts.append(r"\clearpage")
	parts.append(r"\pagestyle{fancy}")
	return parts

def make_author_index_section(author_index: dict[str, list[dict]]) -> list[str]:
	"""
	Author Index in 2 columns.
	Header printed at the top of BOTH columns on EACH page.
	We do manual pagination because multicols cannot repeat headers per column/page automatically.
	"""
	ROWS_PER_COLUMN = 46  # <-- TUNE THIS (try 40–55 depending on font/spacing)

	def header_lines() -> list[str]:
		return [
			r"\parbox[t]{0.13\columnwidth}{\raggedleft\textbf{Initials}} "
			r"\parbox[t]{0.52\columnwidth}{\raggedright\textbf{Surname}} "
			r"\parbox[t]{0.26\columnwidth}{\raggedright\textbf{Pages}}\par",
			r"\hrule\vspace{4pt}",
		]

	parts: list[str] = []
	parts.append(r"\clearpage")
	parts.append(r"\section*{Author Index}")
	parts.append(r"\addcontentsline{toc}{section}{Author Index}")
	parts.append(r"\vspace{0.5em}")

	# ---- Build all rows first (as LaTeX lines) ----
	authors_sorted = sorted(author_index.keys(), key=author_sort_key)
	row_lines: list[str] = []

	for author in authors_sorted:
		items = author_index.get(author, [])
		items_sorted = sorted(items, key=lambda it: it.get("id", ""))

		links = [r"\hyperlink{%s}{\pageref{%s}}" % (it["id"], it["label"]) for it in items_sorted]
		pages_tex = ", ".join(links) if links else ""

		initials, surname = split_author_initials_surname(author)
		initials_tex = latex_escape(initials)
		surname_tex  = latex_escape(surname)

		row_lines.append(
			r"\parbox[t]{0.13\columnwidth}{\raggedleft %s} "
			r"\parbox[t]{0.52\columnwidth}{\raggedright %s} "
			r"\parbox[t]{0.26\columnwidth}{\raggedright %s}\par"
			r"\vspace{3pt}"
			% (initials_tex, surname_tex, pages_tex)
		)

	# ---- Manual pagination ----
	rows_per_page = 2 * ROWS_PER_COLUMN
	n = len(row_lines)

	# styling for the whole index
	parts.append(r"\small")
	parts.append(r"\setlength{\parindent}{0pt}")

	i = 0
	first_page = True
	while i < n:
		if not first_page:
			parts.append(r"\clearpage")
		first_page = False

		left = row_lines[i : i + ROWS_PER_COLUMN]
		right = row_lines[i + ROWS_PER_COLUMN : i + rows_per_page]
		i += rows_per_page
		
		parts.append(r"\thispagestyle{plain}")
		parts.append(r"\begin{multicols}{2}")
		parts.append(r"\setlength{\columnsep}{18pt}")

		# Column 1 header + rows
		parts.extend(header_lines())
		parts.extend(left)

		# Column 2 header + rows
		parts.append(r"\columnbreak")
		parts.extend(header_lines())
		parts.extend(right)

		parts.append(r"\end{multicols}")

	return parts


def build_author_index(records: list[dict]) -> dict[str, list[dict]]:
	"""
	Returns: { "Author Name": [ {id, label, title, area}, ... ], ... }
	"""
	idx: dict[str, list[dict]] = {}
	for r in records:
		for a in r.get("authors", []):
			idx.setdefault(a, []).append({
				"id": r["id"],
				"label": r["label"],
				"title": r["title"],
				"area": r["area"],
			})
	# sort authors alphabetically (case-insensitive)
	sorted_items = sorted(idx.items(), key=lambda kv: kv[0].casefold())
	return dict(sorted(idx.items(), key=lambda kv: author_sort_key(kv[0])))


def latex_escape(text: str) -> str:
		"""Escape LaTeX special characters, but keep normal Unicode."""
		if text is None:
				return ""
		text = text.replace("\r\n", "\n").replace("\r", "\n")
		return "".join(_LATEX_SPECIALS.get(ch, ch) for ch in text)


def make_transition_page(area: str) -> list[str]:
	# A clean title/transition page; no page number shown
	area_esc = latex_escape(area)
	return [
		r"\clearpage",
		r"\thispagestyle{empty}",
		r"\vspace*{\fill}",
		r"\begin{center}",
		r"\Huge " + area_esc,
		r"\end{center}",
		r"\vspace*{\fill}",
		r"\clearpage",
	]

	return



	






def author_sort_key(name: str) -> tuple[str, str]:
	initials, surname = split_author_initials_surname(name)
	return (surname.casefold(), initials.casefold(), name.casefold())


def split_author_initials_surname(name: str) -> tuple[str, str]:
	"""
	Split at the end of the initials (first '.' sequence).
	Examples:
	- 'P. Van Morisson' -> ('P.', 'Van Morisson')
	- 'P.A. Van Morisson' -> ('P.A.', 'Van Morisson')
	- 'P. A. Van Morisson' -> ('P. A.', 'Van Morisson')
	"""
	s = (name or "").strip()
	s = re.sub(r"\s+", " ", s)

	# If there is no dot, fallback: no initials, whole string is surname/name
	if "." not in s:
		return ("", s)

	# Find the position of the LAST dot within the "initials block" at the start.
	# This matches patterns like: "P.", "P.A.", "P. A.", "P. A. B."
	m = re.match(r"^([A-Za-z](?:\.\s*[A-Za-z])*\.)\s*(.*)$", s)
	if m:
		initials = m.group(1).replace(" ", "") if ". " not in m.group(1) else m.group(1)
		rest = m.group(2).strip()
		return (initials.strip(), rest if rest else "")

	# Simpler fallback: split at first '.' only
	i = s.find(".")
	return (s[:i+1].strip(), s[i+1:].strip())








def main() -> None:
	xlsx_path = Path(XLSX_PATH)
	out_tex = Path(OUT_TEX)

	wb = openpyxl.load_workbook(xlsx_path, data_only=True)
	ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

	pdf_dir = Path(r"Z:\ikkv\Dokus_LKKV\040_Projekte\02_Eigenforschung\2025-09-Thomas\Book-of-Abstracts\downloaded_pdfs")
	per_area: dict[str, list[dict]] = {a: [] for a in AREA_ORDER}

	for row in range(START_ROW, ws.max_row + 1):
		
		url_cell = ws.cell(row=row, column=URL_COL)
		url = get_cell_url(url_cell)
		url_text = get_cell_text(url_cell)

		area = get_cell_text(ws.cell(row=row, column=AREA_COL))
		if area not in AREA_ORDER:
			continue

		title = get_cell_text(ws.cell(row=row, column=TITLE_COL))
		authors_cell = get_cell_text(ws.cell(row=row, column=AUTHOR_COL))
		main_author = parse_main_author(authors_cell)
		authors_list = parse_all_authors(authors_cell)


		per_area[area].append({
			"row": row,
			"area": area,
			"title": title if title else "(No title)",
			"main_author": main_author,
			"authors": authors_list,
			"url": url,			  
			"url_text": url_text,	
			})


	records: list[dict] = []
	idx = 0
	
	for area in AREA_ORDER:
		for rec in per_area[area]:
			idx += 1
	
			local_name = filename_for_record(rec, idx)
			pdf_path = pdf_dir / local_name
	
			try:
				if not (pdf_path.exists() and is_pdf_file(pdf_path)):
					if rec.get("url"):
						download_pdf(rec["url"], pdf_path)
						src = "URL"
					else:
						local_pdf = find_local_pdf(rec, idx, url_cell_text=rec.get("url_text", ""))
						if local_pdf is None:
							print(f"NO SOURCE (empty URL + not found locally) row {rec['row']} | title={rec['title']}")
							continue
						pdf_path.parent.mkdir(parents=True, exist_ok=True)
						shutil.copy2(local_pdf, pdf_path)
						src = "LOCAL"
				else:
					src = "CACHE"
	
				if not is_pdf_file(pdf_path):
					print(f"NOT A PDF (skipping) row {rec['row']}: {pdf_path.name}")
					try:
						pdf_path.unlink()
					except Exception:
						pass
					continue
	
				rec2 = dict(rec)
				rec2["pdf_path"] = pdf_path
				rec2["id"] = f"abs:{idx:04d}"
				rec2["label"] = f"lab:{idx:04d}"
				records.append(rec2)
	
				print(f"OK ({src}) row {rec['row']} | area={area} | file={pdf_path.name}")
	
			except Exception as e:
				print(f"ERROR row {rec['row']} area={area} url={rec.get('url','')} reason={e}")


	if not records:
		raise RuntimeError("No valid PDFs downloaded for the requested AREA_ORDER list.")

	build_tex(records, out_tex)
	print(f"PDFs saved under: {pdf_dir}")
	print("Compile from the folder containing the .tex:")
	print(f"  pdflatex {out_tex.name}")
	print("If pdflatex fails due to PDF compatibility, try:")
	print(f"  lualatex {out_tex.name}")



#%% Main
if __name__ == "__main__":
	main()
